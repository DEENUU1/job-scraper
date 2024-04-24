from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from utils.get_current_date import get_current_date
from schemas.offer import Offer
import os


class ExcelWriter:
    """
    Class for writing data to an Excel file.
    """

    def __init__(self, file_name: str = "job_offers.xlsx") -> None:
        """Initialize ExcelWriter with the filename of the Excel file.

        Args:
            file_name (str): The filename of the Excel file.
        """
        self.file_name = file_name
        if os.path.exists(file_name):
            self.workbook = load_workbook(filename=file_name)
        else:
            self.workbook = Workbook()
        self.sheet = self.workbook.active

    def data_exists(self, url: str, url_column: str = "B") -> bool:
        """Check if data exists in the Excel file.

        Args:
            url_column (str): The column letter where URLs are stored.
            url (str): The URL to check for existence.

        Returns:
            bool: True if data exists, False otherwise.
        """
        try:
            for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=True):
                if row[self.column_index(url_column) - 1] == url:
                    return True
            return False

        except Exception as e:
            print(e)
            return False

    @staticmethod
    def column_index(column_letter: str) -> int:
        """Converts column letter into column index.

        Args:
            column_letter (str): The letter representing the column.

        Returns:
            int: The corresponding column index.
        """
        return ord(column_letter.upper()) - 64

    def add_data(self, data: Offer, website: str, tag: Optional[str]) -> None:
        """Add data to the Excel file.

        Args:
            data (Offer): The offer data to add.
            website (str): The website associated with the offer.
            tag (str): The tag associated with the offer.
        """
        next_row = self.sheet.max_row + 1
        row_data = [
            data.title,
            data.url,
            website,
            str(get_current_date()),
            tag
        ]
        for idx, value in enumerate(row_data, start=1):
            self.sheet.cell(row=next_row, column=idx, value=value)

    def save(self) -> None:
        """Save the Excel file."""
        for column_cells in self.sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            self.sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True)

        self.workbook.save(filename=self.file_name)
        print("Data saved to Excel file")
